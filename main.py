from flask import Flask, render_template, request
import openpyxl

app = Flask(__name__)

# Global değişken oluştur
gonderildi = False


@app.route('/', methods=['GET', 'POST'])
def form():
    global gonderildi

    if request.method == 'POST' and not gonderildi:
        isim = request.form['isim']
        soyisim = request.form['soyisim']
        telefon = request.form['telefon']
        mail = request.form['mail']
        pdf = request.files['pdf']

        # Form verilerini Excel dosyasına kaydetme
        wb = openpyxl.Workbook()
        ws = wb.active

        ws['A1'] = 'İsim'
        ws['B1'] = 'Soyisim'
        ws['C1'] = 'Telefon'
        ws['D1'] = 'E-posta'

        ws['A2'] = isim
        ws['B2'] = soyisim
        ws['C2'] = telefon
        ws['D2'] = mail

        wb.save('veriler.xlsx')

        # PDF dosyasını kaydetme
        pdf.save('uploads/' + pdf.filename)

        gonderildi = True

        return 'Form başarıyla gönderildi!'

    return render_template('form.html')


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080, debug=True)
